import openpyxl


def load_excel():
    global wb
    global sh
    wb = openpyxl.load_workbook("data/data.xlsx")
    sh = wb.active


def get_cell_data(rowNo, colNo):
    return sh.cell(rowNo, colNo).value


def save_data():
    wb.save("data_result.xlsx")


def get_data_as_list_tuples(n):
    for i in range(1, sh.max_row + 1):
       return sh.cell(row=i, column=n).value


"""
def get_max_rows():


def get_max_columns():
"""
