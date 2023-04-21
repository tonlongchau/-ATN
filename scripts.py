import openpyxl
from selenium.webdriver import Keys, ActionChains
from selenium import webdriver
from selenium.webdriver.common.by import By
import time


path = "data/testdata1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
nrows = sheet_obj.max_row
ncol = sheet_obj.max_column


for i in range(2, nrows+1):
    Module = sheet_obj.cell(row=i, column=1).value
    TestCase = sheet_obj.cell(row=i, column=2).value
    xpath = sheet_obj.cell(row=i, column=3).value
    action_type = sheet_obj.cell(row=i, column=4).value
    data = sheet_obj.cell(row=i, column=5).value
    text = sheet_obj.cell(row=i, column=6).value
    
    if data == "Chrome":
        try:
            driver = webdriver.Chrome(executable_path='./Report-Final/webdriver/chromedriver')
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            sheet_obj.cell(row=i, column=6).value = "FAIL"
    if data == "Firefox":
        try:
            driver = webdriver.Firefox(executable_path='./Report-Final/webdriver/geckodriver')
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            sheet_obj.cell(row=i, column=6).value = "FAIL"
            
    actions = ActionChains(webdriver)
    if Module == "Navigate":
        try:
            driver.get(data)
            sheet_obj.cell(row=i, column=6).value = "PASS"
            driver.maximize_window()
        except Exception:
            sheet_obj.cell(row=i, column=6).value = "FAIL"
            
    if action_type == "textbox":
        try:
            print(xpath)
            driver.find_element(By.XPATH, xpath).send_keys(data)
            print("Entered Data")
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            print("Unable to Enter Data")
            sheet_obj.cell(row=i, column=6).value = "FAIL"

            
    if action_type == "click":
        try:
            time.sleep(3)
            driver.find_element(By.XPATH, xpath).click()
            print("Clicked Successfully")
            time.sleep(1)
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            print("Unable to Click")
            sheet_obj.cell(row=i, column=6).value = "FAIL"

    if action_type == "select":
        try:
            print(xpath)
            element = driver.find_element(By.XPATH, xpath)
            element.is_enabled()
            element.send_keys(Keys.CONTROL + "a")
            element.send_keys(text)
            from keyboard import press
            press('enter')
            print("Entered Data")
            sheet_obj.cell(row=i, column=6).value = "PASS"

        except Exception:
            print("Unable to Enter Data")
            sheet_obj.cell(row=i, column=6).value = "FAIL"

    if action_type == "autocomplete":
        try:
            print(xpath)
            element = driver.find_element(By.XPATH, xpath)
            element.send_keys(text)
            element.send_keys(Keys.ARROW_DOWN)
            element.send_keys(Keys.ENTER)
            time.sleep(3)
            print("Entered Data")
            sheet_obj.cell(row=i, column=6).value = "PASS"

        except Exception:
            print("Unable to Enter Data")
            sheet_obj.cell(row=i, column=6).value = "FAIL"



wb_obj.save("data.xlsx")

time.sleep(3)
driver.save_screenshot("/home/anhtan/Report-Selenium/test.png")

driver.quit()